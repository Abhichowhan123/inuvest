import os
import pandas as pd
from django.shortcuts import render
from django.core.files.storage import FileSystemStorage
from openpyxl.utils import get_column_letter
import openpyxl


def upload(request):
    if request.method == 'POST':
        df = pd.read_csv(request.FILES['document'])
        df.to_excel('my_file.xlsx', index=None, header=True)
        wb = openpyxl.load_workbook('my_file.xlsx')
        # wb = openpyxl.load_workbook(request.FILES['document'])
        ws = wb.active
        # print(ws['B3'].value )
        # ws['B3'].value = 9999999
        A = ["TR", "+DM 1", "-DM 1", "TR14", "+DM14", "-DM14", "+DI14", "-DI14", "DI 14 Diff", "DI 14 Sum", "DX", "ADX"]
        w, e = 6, 0
        for y in range(0, 12):
            ws[get_column_letter(w) + str(1)].value = A[y]
            w += 1
        result, i = 0, 3
        j = 5
        def fun(a):
            result, i = 0, 2
            for row in ws:
                if i == 17:
                    break
                else:
                    if ws[get_column_letter(a) + str(i)].value == None:
                        result += 0
                    else:
                        result += ws[get_column_letter(a) + str(i)].value
                i += 1
            return result

        def adx():
            result, i = 0, 16
            for row in ws:
                if i == 30:
                    break
                else:
                    if ws[get_column_letter(16) + str(i)].value == None:
                        result += 0
                    else:
                        result += ws[get_column_letter(16) + str(i)].value
                i += 1
            return result / 14

        def calculation(k, l):
            return ws[get_column_letter(k) + str((i) - 1)].value - (
                    ws[get_column_letter(k) + str((i) - 1)].value / 14) + ws[get_column_letter(l) + str(i)].value

        for row in ws:
            high = ws[get_column_letter(3) + str(i)].value
            low = ws[get_column_letter(4) + str(i)].value
            close = ws[get_column_letter(5) + str(i)].value
            if high != None and low != None and close != None:
                TR = max((high - low), abs(high - close), abs(low - close))
                ws[get_column_letter(j + 1) + str(i)].value = TR

                # +DM1
                if ((high - ws[get_column_letter(3) + str((i) - 1)].value) > (
                        ws[get_column_letter(4) + str((i) - 1)].value - low)):
                    DM1 = max((high - ws[get_column_letter(3) + str((i) - 1)].value), 0)
                else:
                    DM1 = 0
                ws[get_column_letter(j + 2) + str(i)].value = DM1

                # -DM1
                if (((ws[get_column_letter(4) + str((i) - 1)].value) - low) > (
                        high - ws[get_column_letter(3) + str((i) - 1)].value)):
                    DM12 = max(((ws[get_column_letter(4) + str((i) - 1)].value) - low), 0)
                else:
                    DM12 = 0
                ws[get_column_letter(j + 3) + str(i)].value = round(DM12, 2)
                if i == 16:
                    # TR14
                    ws[get_column_letter(j + 4) + str(i)].value = round(fun(6), 2)
                    # +DM14
                    ws[get_column_letter(j + 5) + str(i)].value = round(fun(7), 2)
                    # -DM14
                    ws[get_column_letter(j + 6) + str(i)].value = round(fun(8), 2)

                    # +DI14
                    DI14 = 100 * (
                                ws[get_column_letter(10) + str((i))].value / ws[get_column_letter(9) + str((i))].value)
                    ws[get_column_letter(j + 7) + str(i)].value = round(DI14, 2)
                    # -DI14
                    DI14_second = 100 * (
                            ws[get_column_letter(11) + str((i))].value / ws[get_column_letter(9) + str((i))].value)
                    ws[get_column_letter(j + 8) + str(i)].value = round(DI14_second, 2)
                    # DI 14 Diff
                    DI_14_Diff = abs(
                        ws[get_column_letter(12) + str((i))].value - ws[get_column_letter(13) + str((i))].value)
                    ws[get_column_letter(j + 9) + str(i)].value = round(DI_14_Diff, 2)

                    # DI 14 SUM
                    DI_14_SUM = ws[get_column_letter(12) + str((i))].value + ws[get_column_letter(13) + str((i))].value
                    ws[get_column_letter(j + 10) + str(i)].value = round(DI_14_SUM, 2)
                    # DX
                    DX = 100 * (ws[get_column_letter(14) + str((i))].value / ws[get_column_letter(15) + str((i))].value)
                    ws[get_column_letter(j + 11) + str(i)].value = round(DX, 2)
                if i > 16:
                    # TR14

                    TR14 = calculation(9, 6)
                    ws[get_column_letter(j + 4) + str(i)].value = round(TR14, 2)
                    # +DM14
                    DM14 = calculation(10, 7)
                    ws[get_column_letter(j + 5) + str(i)].value = round(DM14, 2)
                    # -DM14
                    DM14_second = calculation(11, 8)
                    ws[get_column_letter(j + 6) + str(i)].value = round(DM14_second, 2)

                    # +DI14
                    DI14 = 100 * (ws[get_column_letter(10) + str(i)].value / ws[get_column_letter(9) + str((i))].value)
                    ws[get_column_letter(j + 7) + str(i)].value = round(DI14, 2)
                    # -DI14
                    DI14_second = 100 * (
                            ws[get_column_letter(11) + str((i))].value / ws[get_column_letter(9) + str((i))].value)
                    ws[get_column_letter(j + 8) + str(i)].value = round(DI14_second, 2)
                    # DI 14 Diff
                    DI_14_Diff = abs(
                        ws[get_column_letter(12) + str((i))].value - ws[get_column_letter(13) + str((i))].value)
                    ws[get_column_letter(j + 9) + str(i)].value = round(DI_14_Diff, 2)

                    # DI 14 SUM
                    DI_14_SUM = ws[get_column_letter(12) + str((i))].value + ws[get_column_letter(13) + str((i))].value
                    ws[get_column_letter(j + 10) + str(i)].value = round(DI_14_SUM, 2)
                    # DX
                    DX = 100 * (ws[get_column_letter(14) + str((i))].value / ws[get_column_letter(15) + str((i))].value)
                    ws[get_column_letter(j + 11) + str(i)].value = round(DX, 2)
                if i == 29:
                    w = adx()
                    ws[get_column_letter(j + 12) + str(i)].value = round(w, 2)
                if i > 29:
                    ADX = (ws[get_column_letter(17) + str((i) - 1)].value * 13 + ws[
                        get_column_letter(16) + str((i))].value) / 14
                    ws[get_column_letter(j + 12) + str(i)].value = round(ADX, 2)
            i += 1

        # wb.save('my_file.xlsx')
        # wb = openpyxl.load_workbook(request.FILES['document'])
        # ws = wb.active
        # print(ws)
        # uploaded_file = request.FILES['document']
        #
        # fs = FileSystemStorage()
        # fs.save(uploaded_file.name, uploaded_file)


    return render(request,'uplode.html')

def Visualization(request):
    return render(request,'Visualization.html')
